import openpyxl
from openpyxl import load_workbook
import unicodedata
import pandas as pd
import re

#Read datasheets
sample_wb = load_workbook("/Users/noravandevoorde/Downloads/SPHERE/Add_Population_copy.xlsx")
sample_ws = sample_wb.active
census_wb = load_workbook("/Users/noravandevoorde/Downloads/SPHERE/tabela4714.xlsx")
census_ws = census_wb.active

#Print column names in both datasets
print("Columns in sample datasheet:", [cell.value for cell in sample_ws[1] if cell.value is not None])
print("Columns in census dataset:", [cell.value for cell in census_ws[1] if cell.value is not None])

#remove accents
def remove_accents(input_str):
    if input_str is None:
        return None
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    return ''.join([c for c in nfkd_form if not unicodedata.combining(c)])

#split city and state in census data
df = pd.read_excel("/Users/noravandevoorde/Downloads/SPHERE/tabela4714.xlsx")
df[["City", "State"]] = df["Município"].str.extract(r"^(.*?)\s*\((.*?)\)$")
df["City"] = df["City"].str.strip()
df["State"] = df["State"].str.strip()



#clean names
def clean(value):
    if value is None:
        return None
    return remove_accents(str(value).lower().strip())

#Get cities in sample and clean names
sample_cities = {}
for excel_row, row in enumerate(sample_ws.iter_rows(min_row=2, values_only=True)):
    sample_cities[excel_row] = {
        "City": clean(row[2]) if row[2] is not None else "",
        "State": clean(row[1]) if row [1] is not None else ""
    }

print(list(sample_cities.items())[:5])

#Get cities and population from census data
cities_and_pop = {}
for census_row, row in enumerate(census_ws.iter_rows(min_row=3, values_only=True)):
    location = clean(row[2]) if row[2] is not None else ""
    
    # split "City (ST)" format
    match = re.match(r"^(.*?)\s*\((.*?)\)$", location)

    if match:
        city = clean(match.group(1))
        state = clean(match.group(2))
    else:
        city = location
        state = ""

    cities_and_pop[census_row] = {
        "City": city,
        "State": state,
        "Population": clean(row[3]) if row[3] is not None else ""
    }

print(list(cities_and_pop.items())[:5])

#Match enteries based on city and state and paste population into sample datasheet
#Make a look up table for each city, state and corresponding population
pop_lookup = {}
for _, cities_and_pop_data in cities_and_pop.items():
    key = (cities_and_pop_data["City"], cities_and_pop_data["State"])
    pop_lookup[key] = cities_and_pop_data["Population"]

#Create new population columns
population_col = sample_ws.max_column+1
sample_ws.cell(row=1, column=population_col).value = "Population"

#Match city and state pairs across both worksheets and paste in the population
city_state_matches = 0

for row in range(2, sample_ws.max_row + 1):
    state = clean(sample_ws.cell(row=row, column=2).value)  # State in column B
    city = clean(sample_ws.cell(row=row, column=3).value)   # City in column C

    key = (city, state)
    current_label = sample_ws.cell(row=row, column=20).value

    if (not current_label or str(current_label).strip() == "") and (key in pop_lookup):
        sample_ws.cell(row=row, column=20).value = pop_lookup[key]  
        city_state_matches += 1

print(f"Number of city-state matches: {city_state_matches}")
sample_wb.save("Add_Population_copy.xlsx")